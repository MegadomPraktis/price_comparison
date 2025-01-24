import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.base import MIMEBase
from email import encoders
import os
import time
import random
import pandas as pd
import requests
from bs4 import BeautifulSoup
from concurrent.futures import ThreadPoolExecutor, as_completed
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment
from xlsxwriter.utility import xl_rowcol_to_cell
from openpyxl.styles import Font, PatternFill
from datetime import datetime

# Constants for search URLs
PRAKTIS_SEARCH_URL = "https://praktis.bg/catalogsearch/result/?q={}"
PRAKTIKER_SEARCH_URL = "https://praktiker.bg/search/{}"

# Initialize session and headers
session = requests.Session()
USER_AGENTS = [
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/114.0.0.0 Safari/537.36",
    "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.124 Safari/537.36",
    "Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.124 Safari/537.36",
]
session.headers.update({"Accept-Language": "en-US,en;q=0.9"})

def send_email(smtp_server, port, sender_email, sender_password, recipient_emails, subject, body, attachment_path):
    try:
        # Create email message
        msg = MIMEMultipart()
        msg['From'] = sender_email
        msg['To'] = ", ".join(recipient_emails)  # Join multiple recipients with a comma
        msg['Subject'] = subject

        # Attach email body
        msg.attach(MIMEText(body, 'plain'))

        # Attach the file
        with open(attachment_path, "rb") as attachment:
            part = MIMEBase('application', 'octet-stream')
            part.set_payload(attachment.read())
        encoders.encode_base64(part)
        part.add_header('Content-Disposition', f'attachment; filename="{os.path.basename(attachment_path)}"')
        msg.attach(part)

        # Connect to the SMTP server using implicit SSL/TLS
        with smtplib.SMTP_SSL(smtp_server, port) as server:
            server.login(sender_email, sender_password)
            server.send_message(msg)

        print("Email sent successfully to:", ", ".join(recipient_emails))
    except Exception as e:
        print(f"Failed to send email: {e}")


def get_soup(url):
    for attempt in range(3):  # Retry up to 3 times
        try:
            session.headers.update({"User-Agent": random.choice(USER_AGENTS)})
            response = session.get(url, timeout=16)
            response.raise_for_status()
            return BeautifulSoup(response.content, 'html.parser')
        except requests.RequestException:
            time.sleep(2 ** attempt + random.uniform(0.5, 1.5))  # Exponential backoff
    return None

def fetch_product_data_praktis(code):
    code = str(code).strip()
    url = PRAKTIS_SEARCH_URL.format(code)
    soup = get_soup(url)
    if not soup:
        return {"code": code, "name": "N/A", "url": url, "regular_price": "N/A", "promo_price": "N/A"}

    name = soup.select_one("p.product-name.h4")
    regular_price = soup.select_one("span.price.striked, div.old-price span.price") or soup.select_one("span.price")
    promo_price = soup.select_one("div.special-price span.price")

    return {
        "code": code,
        "name": name.text.strip() if name else "N/A",
        "url": url,
        "regular_price": regular_price.text.strip().replace("\u043b\u0432.", "").strip() if regular_price else "N/A",
        "promo_price": promo_price.text.strip().replace("\u043b\u0432.", "").strip() if promo_price else None,
    }

def fetch_product_data_praktiker(code):
    code = str(code).strip()
    url = PRAKTIKER_SEARCH_URL.format(code)
    soup = get_soup(url)
    if not soup:
        return {"code": code, "name": "N/A", "url": url, "regular_price": None, "promo_price": None}

    name_element = soup.select_one("h2.product-item__title a")
    name = name_element.text.strip() if name_element else "N/A"

    regular_price = None
    promo_price = None

    old_price_element = soup.select_one("span.product-price--old .product-price__value")
    old_price_sup = old_price_element.find_next("sup") if old_price_element else None
    if old_price_element:
        regular_price = old_price_element.text.strip()
        if old_price_sup:
            regular_price += "." + old_price_sup.text.strip()
    else:
        regular_price_element = soup.select_one("span.product-price__value, span.price__value")
        regular_price_sup = regular_price_element.find_next("sup") if regular_price_element else None
        if regular_price_element:
            regular_price = regular_price_element.text.strip()
            if regular_price_sup:
                regular_price += "." + regular_price_sup.text.strip()

    promo_price_element = soup.select_one(
        "div.product-store-prices__item > span.product-price:not(.product-price--old) span.product-price__value"
    )
    promo_price_sup = promo_price_element.find_next("sup") if promo_price_element else None
    if promo_price_element:
        promo_price = promo_price_element.text.strip()
        if promo_price_sup:
            promo_price += "." + promo_price_sup.text.strip()

    return {
        "code": code,
        "name": name,
        "url": url,
        "regular_price": regular_price,
        "promo_price": promo_price,
    }

def adjust_excel_formatting(file_path):
    workbook = load_workbook(file_path)
    sheet = workbook.active

    for col in sheet.columns:
        max_length = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
                cell.alignment = Alignment(wrap_text=True)
            except Exception as e:
                print(f"Error adjusting column {col_letter}: {e}")
        adjusted_width = max_length + 2
        sheet.column_dimensions[col_letter].width = adjusted_width

    workbook.save(file_path)
    workbook.close()

def process_excel_and_split_files(input_file, output_file_1, output_file_2):
    try:
        start_time = time.time()

        df = pd.read_excel(input_file, engine="odf")
        pairs = df.values.tolist()

        results = []
        with ThreadPoolExecutor(max_workers=5) as executor:
            futures = {
                executor.submit(fetch_product_data_praktis, pair[0]): pair for pair in pairs
            }

            for future in as_completed(futures):
                pair = futures[future]
                praktis_data = future.result()
                praktiker_data = fetch_product_data_praktiker(pair[1])
                results.append({
                    "Praktis Code": pair[0],
                    "Praktiker Code": pair[1],
                    "Praktis Name": praktis_data["name"],
                    "Praktiker Name": praktiker_data["name"],
                    "Praktis Regular Price": praktis_data["regular_price"],
                    "Praktiker Regular Price": praktiker_data["regular_price"],
                    "Praktis Promo Price": praktis_data["promo_price"],
                    "Praktiker Promo Price": praktiker_data["promo_price"],
                })

        # Split results into two parts
        results_part_1 = results[:180]
        results_part_2 = results[180:]

        # Helper function to write data to Excel
        def write_to_excel(file_path, data):
            with pd.ExcelWriter(file_path, engine="xlsxwriter") as writer:
                output_df = pd.DataFrame(data)
                output_df.to_excel(writer, index=False, sheet_name="Product Details")

                workbook = writer.book
                worksheet = writer.sheets["Product Details"]

                # Add hyperlinks and adjust formatting
                for row_num, row_data in enumerate(data, start=1):
                    if row_data["Praktis Name"]:
                        praktis_url = PRAKTIS_SEARCH_URL.format(row_data["Praktis Code"])
                        worksheet.write_url(
                            row_num, 2, praktis_url, string=row_data["Praktis Name"]
                        )
                    if row_data["Praktiker Name"]:
                        praktiker_url = PRAKTIKER_SEARCH_URL.format(row_data["Praktiker Code"])
                        worksheet.write_url(
                            row_num, 3, praktiker_url, string=row_data["Praktiker Name"]
                        )

                # Adjust column widths and enable text wrapping
                for col_num, col_data in enumerate(output_df.columns):
                    max_length = max([len(str(val)) for val in output_df[col_data].fillna("")] + [len(col_data)])
                    worksheet.set_column(col_num, col_num, max_length + 2, writer.book.add_format({'text_wrap': True}))

        # Write data to two files
        write_to_excel(output_file_1, results_part_1)
        write_to_excel(output_file_2, results_part_2)

        end_time = time.time()
        print(f"Data exported successfully to {output_file_1} and {output_file_2}")
        print(f"Execution time: {end_time - start_time:.2f} seconds")

    except Exception as e:
        print(f"An error occurred: {e}")


if __name__ == "__main__":
    input_excel = r"C:\Users\angel\PycharmProjects\price_comparison_files\Input_files\cleaned_products_list_test.ods"
    timestamp = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
    output_excel_1 = rf"C:\Users\angel\PycharmProjects\price_comparison_files\Output_files\product_details_part1_{timestamp}.xlsx"
    output_excel_2 = rf"C:\Users\angel\PycharmProjects\price_comparison_files\Output_files\product_details_part2_{timestamp}.xlsx"

    process_excel_and_split_files(input_excel, output_excel_1, output_excel_2)

    smtp_server = "mail.praktis.bg"
    port = 465
    sender_email = "a.borisov@praktis.bg"
    sender_password = ""

    # Email for the first file
    recipient_emails_1 = ["b.borisov@praktis.bg"]
    subject_1 = "Репорт за сравнение на цените - Боби"
    body_1 = f"Здравей,\n\nТова е тестови файл за сравнение на цени, генериран на: {timestamp}.\n\nПоздрави,\nСкрипт"
    send_email(smtp_server, port, sender_email, sender_password, recipient_emails_1, subject_1, body_1, output_excel_1)

    # Email for the second file
    recipient_emails_2 = ["a.borisov@praktis.bg"]
    subject_2 = "Репорт за сравнение на цените - Ангел"
    body_2 = f"Здравей,\n\nТова е тестови файл за сравнение на цени, генериран на: {timestamp}.\n\nПоздрави,\nСкрипт"
    send_email(smtp_server, port, sender_email, sender_password, recipient_emails_2, subject_2, body_2, output_excel_2)
